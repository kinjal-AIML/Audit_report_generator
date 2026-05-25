"""
PANCHOT Branch Report Extractor → Excel Master Sheet
=====================================================
Extracts structured data from 3 bank branch PDFs and saves into
a formatted multi-sheet Excel workbook.

Sheets produced:
  1. Cash Summary          ← PANCHOT_BR_CASH_FEB_SUM.pdf
  2. Insurance Pending     ← PANCHOT_BR_JAN_2026_INSURANCE_PENDING_REGI_.pdf
  3. Overdue Report        ← PANCHOT_BR_JAN_2026_OVER_DUE_REPORT.pdf

Requirements:
    pip install pdfplumber openpyxl

Usage:
    python panchot_report_extractor.py
"""

import os
import re
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# CHAR-LEVEL WORD RECONSTRUCTION
# These PDFs use a spaced-character font rendering where individual letters
# appear with 0.01–0.11 px gaps within a visual word, and ~5.5 px gaps
# between visual words. Standard pdfplumber x_tolerance cannot handle this.
# ─────────────────────────────────────────────────────────────────────────────

INTRA_WORD_GAP_PX = 3.0    # gap within same visual word token
INTER_WORD_GAP_PX = 15.0   # gap above which = field/column boundary


def _chars_to_words(chars):
    """
    Group page chars into word tokens using pixel-gap thresholds.

    Returns: list of (x0, x1, text) tuples representing distinct tokens.
    Gaps ≤ INTRA_WORD_GAP_PX  → same token (chars run together)
    Gaps 3–15 px              → word boundary (space between words)
    Gaps > 15 px              → column/field boundary
    """
    if not chars:
        return []
    chars_sorted = sorted(chars, key=lambda c: c["x0"])
    tokens = []
    buf = [chars_sorted[0]]

    for c in chars_sorted[1:]:
        gap = c["x0"] - buf[-1]["x1"]
        if gap <= INTRA_WORD_GAP_PX:
            buf.append(c)
        else:
            tokens.append((buf[0]["x0"], buf[-1]["x1"], "".join(ch["text"] for ch in buf)))
            buf = [c]
    tokens.append((buf[0]["x0"], buf[-1]["x1"], "".join(ch["text"] for ch in buf)))
    return tokens


def _page_rows(page, y_group_tolerance=3):
    """
    Return ordered list of rows, each row = list of (x0, x1, text) tokens,
    grouped by y-position with a tolerance.
    """
    chars = [c for c in page.chars if c["text"].strip()]
    row_map = {}
    for c in chars:
        key = round(c["top"] / y_group_tolerance) * y_group_tolerance
        row_map.setdefault(key, []).append(c)

    rows = []
    for top in sorted(row_map):
        tokens = _chars_to_words(row_map[top])
        if tokens:
            rows.append((top, tokens))
    return rows


def _all_rows_multipage(pdf_path):
    """Return (page_idx, top, tokens) for all pages."""
    result = []
    with pdfplumber.open(pdf_path) as pdf:
        for pi, page in enumerate(pdf.pages):
            for top, tokens in _page_rows(page):
                result.append((pi, top, tokens))
    return result


# ─────────────────────────────────────────────────────────────────────────────
# NOISE FILTER
# ─────────────────────────────────────────────────────────────────────────────

_NOISE = re.compile(
    r"THE MEHSANA|PANCHOT BRANCH|MICR:-|IFSC:-|Print Date|"
    r"^-{5,}$|User Name:|Page \d+ of|^Clerk\s|^Cashier\s|"
    r"Security Details|A/c No\.\s+Name\s+NPA|"
    r"Overdue Report For All|Cash Summary Report|Insurance Pending Register",
    re.IGNORECASE,
)


def _is_noise(text):
    return bool(_NOISE.search(text))


def _token_text(tokens):
    return " ".join(t[2] for t in tokens)


# ─────────────────────────────────────────────────────────────────────────────
# 1. CASH SUMMARY
# ─────────────────────────────────────────────────────────────────────────────

def extract_cash_summary(pdf_path):
    """
    Parses Cash Summary Report.
    Returns DataFrame with columns:
      Section | Description | Count | Amount (INR)
    """
    rows = []
    report_date = ""

    all_rows = _all_rows_multipage(pdf_path)

    for _, _, tokens in all_rows:
        line = _token_text(tokens)
        if _is_noise(line):
            continue

        # Report date
        m = re.search(r"AsOn\s+(\S+)", line, re.IGNORECASE)
        if m:
            report_date = m.group(1)
            continue

        # Cash flow lines: "Opening Cash Balance  858838.00"
        m = re.match(
            r"^(Opening Cash Balance|Total Cash Receipt|Total Cash Payment|Closing Cash Balance)\s+([\d,\.]+)$",
            line,
        )
        if m:
            rows.append({"Section": "Cash Flow", "Description": m.group(1),
                         "Count": "", "Amount (INR)": m.group(2)})
            continue

        # Denomination lines — split by x-position
        # Tokens on denom rows: [denomination_label, count, amount]
        # e.g. ['N500', 'NOTE', '1580', '790000.00']
        if len(tokens) >= 3:
            last_tok = tokens[-1][2]
            second_last = tokens[-2][2]
            # Amount is last token (float), count is second-last (int)
            if re.match(r"^[\d,]+\.?\d*$", last_tok) and re.match(r"^\d+$", second_last):
                label = " ".join(t[2] for t in tokens[:-2])
                if any(kw in label.upper() for kw in ["NOTE", "COIN", "PAISA"]):
                    rows.append({
                        "Section": "Denomination Breakup",
                        "Description": label,
                        "Count": second_last,
                        "Amount (INR)": last_tok,
                    })
                    continue

        # Total
        m = re.match(r"^Total:\s*([\d,\.]+)$", line)
        if m:
            rows.append({"Section": "Denomination Breakup", "Description": "TOTAL",
                         "Count": "", "Amount (INR)": m.group(1)})
            continue

        # Retention limit
        m = re.search(r"Branch Cash Retention Limit\s*:\s*([\d,\.]+)", line)
        if m:
            rows.append({"Section": "Info", "Description": "Branch Cash Retention Limit",
                         "Count": "", "Amount (INR)": m.group(1)})

    df = pd.DataFrame(rows, columns=["Section", "Description", "Count", "Amount (INR)"])
    df.attrs["report_date"] = report_date
    return df


# ─────────────────────────────────────────────────────────────────────────────
# 2. INSURANCE PENDING REGISTER
# ─────────────────────────────────────────────────────────────────────────────

_NPA_CODES = re.compile(r"^(SMA\d|0[0-9]|\d{2}|NPA)$")


def extract_insurance_pending(pdf_path):
    """
    Parses Insurance Pending Register.
    Columns: Branch Code | Branch Name | Loan Type Code | Loan Type Name |
             A/c No. | Name | NPA Code | Limit Amount (INR) | Remarks
    """
    rows = []
    report_date = ""
    branch_code = branch_name = loan_type_code = loan_type_name = ""

    all_rows = _all_rows_multipage(pdf_path)

    for _, _, tokens in all_rows:
        line = _token_text(tokens)
        if not line.strip() or _is_noise(line):
            continue

        # Report date
        m = re.search(r"As On\s*:\s*(\S+)", line, re.IGNORECASE)
        if m:
            report_date = m.group(1)
            continue

        # Branch / loan type headers: "(004) - PANCHOT BRANCH" or "(304) - MCC"
        m = re.match(r"^\((\d+)\)\s*-\s*(.+)$", line)
        if m:
            code, name = m.group(1).strip(), m.group(2).strip()
            if code == "004":
                branch_code, branch_name = code, name
            else:
                loan_type_code, loan_type_name = code, name
            continue

        # Summary rows
        m = re.search(
            r"(Type Wise Total|Branch Wise Total|Grand Total).+?:\s*(\d+)\s+([\d,\.]+)",
            line, re.IGNORECASE,
        )
        if m:
            rows.append({
                "Branch Code": branch_code, "Branch Name": branch_name,
                "Loan Type Code": "", "Loan Type Name": "",
                "A/c No.": "", "Name": m.group(1).strip(),
                "NPA Code": "", "Limit Amount (INR)": m.group(3),
                "Remarks": f"Count: {m.group(2)}",
            })
            continue

        # Account rows: first token = 6-digit account number
        if not tokens:
            continue
        acct_tok = tokens[0][2]
        if not re.match(r"^\d{6}$", acct_tok):
            continue

        # Amount is always last token (float)
        amt_tok = tokens[-1][2]
        if not re.match(r"^[\d,]+\.?\d*$", amt_tok):
            continue

        # NPA code: search from end (before amount, could be 2nd-last or earlier)
        npa_code = ""
        name_tokens = []
        for i in range(len(tokens) - 2, 0, -1):
            if _NPA_CODES.match(tokens[i][2]):
                npa_code = tokens[i][2]
                name_tokens = tokens[1:i]
                break
        else:
            name_tokens = tokens[1:-1]

        name = " ".join(t[2] for t in name_tokens)

        rows.append({
            "Branch Code": branch_code, "Branch Name": branch_name,
            "Loan Type Code": loan_type_code, "Loan Type Name": loan_type_name,
            "A/c No.": acct_tok, "Name": name,
            "NPA Code": npa_code, "Limit Amount (INR)": amt_tok,
            "Remarks": "",
        })

    df = pd.DataFrame(rows, columns=[
        "Branch Code", "Branch Name", "Loan Type Code", "Loan Type Name",
        "A/c No.", "Name", "NPA Code", "Limit Amount (INR)", "Remarks",
    ])
    df.attrs["report_date"] = report_date
    return df


# ─────────────────────────────────────────────────────────────────────────────
# 3. OVERDUE REPORT
# ─────────────────────────────────────────────────────────────────────────────

_DATE_PAT = re.compile(r"^\d{2}/\d{2}/\d{4}$")
_FLOAT_PAT = re.compile(r"^-?[\d,]+\.?\d*$")
_INT_PAT = re.compile(r"^\d+$")
_CONTACT_PAT = re.compile(r"^\d{9,10}$")


def extract_overdue_report(pdf_path):
    """
    Parses Overdue Report (2-page PDF).
    Page 1: Main account fields (account no., name, amounts, dates, NPA, contact)
    Page 2: IR Balance, Security Amount, Customer ID (same row order as page 1)
    Returns merged DataFrame.
    """
    report_date = ""

    # ── Page 1 ──────────────────────────────────────────────────────────────
    acct_rows = []       # final account rows
    summary_rows = []    # typewise / branchwise totals
    loan_type_code = loan_type_name = ""

    # ── Step 1: Segment rows into blocks ─────────────────────────────────────
    # Each block = [primary_row_tokens, secondary_row_1_tokens, ...]
    # A new block starts when the first token is a 6-digit account number
    # or a summary keyword. Between blocks are noise/header rows.

    with pdfplumber.open(pdf_path) as pdf:
        p1_char_rows = _page_rows(pdf.pages[0])

    _SUMMARY_PAT = re.compile(
        r"^(Typewise|Branchwise|Grand)", re.IGNORECASE
    )

    blocks = []   # list of lists-of-rows, each list is one account block
    current_block = []

    for top, tokens in p1_char_rows:
        line = _token_text(tokens)
        if not line.strip():
            continue

        # Report date
        m = re.search(r"As On\s+(\d{2}/\d{2}/\d{4})", line, re.IGNORECASE)
        if m and not report_date:
            report_date = m.group(1)
            continue

        if _is_noise(line):
            continue

        # Loan type / branch headers
        m = re.match(r"^\((\d+\s*)\)\s*-\s*(.+)$", line)
        if m:
            code, name = m.group(1).strip(), m.group(2).strip()
            if code != "004":
                loan_type_code, loan_type_name = code, name
            continue

        # New account row or summary row → starts a new block
        first_tok = tokens[0][2] if tokens else ""
        is_new_acct = re.match(r"^\d{6}$", first_tok)
        is_summary = _SUMMARY_PAT.match(first_tok)

        if is_new_acct or is_summary:
            if current_block:
                blocks.append(current_block)
            current_block = [(loan_type_code, loan_type_name, top, tokens)]
        else:
            if current_block:
                current_block.append((loan_type_code, loan_type_name, top, tokens))

    if current_block:
        blocks.append(current_block)

    # ── Step 2: Parse each block ──────────────────────────────────────────────
    for block in blocks:
        if not block:
            continue

        ltc, ltn, _, primary_tokens = block[0]
        primary_tok_texts = [(t[0], t[2]) for t in primary_tokens]
        first_tok = primary_tokens[0][2] if primary_tokens else ""

        # ── Summary block ──
        if _SUMMARY_PAT.match(first_tok):
            line = _token_text(primary_tokens)
            nums = [v for _, v in primary_tok_texts if _FLOAT_PAT.match(v)]
            ints = [v for _, v in primary_tok_texts if _INT_PAT.match(v) and v not in nums]
            label = re.match(r"^([\w\s]+Total)", line, re.IGNORECASE)
            summary_rows.append({
                "Loan Type Code": "", "Loan Type Name": "",
                "A/c No.": label.group(1).strip() if label else "Total",
                "Name": f"({ints[0]} accounts)" if ints else "",
                "NPA Code": "", "Category": "",
                "Sanctioned Amt (INR)": nums[0] if len(nums) > 0 else "",
                "Op. Date": "", "Due Date": "", "Int. Rate (%)": "",
                "Adv. Recovery (INR)": nums[1] if len(nums) > 1 else "",
                "Installment Amt (INR)": nums[2] if len(nums) > 2 else "",
                "Outstanding (INR)": nums[3] if len(nums) > 3 else "",
                "Overdue Amt (INR)": nums[4] if len(nums) > 4 else "",
                "Pending Inst.": "", "Overdue Period": "",
                "IR Balance (INR)": nums[5] if len(nums) > 5 else "",
                "Security Amt (INR)": "", "Contact No.": "", "Customer Id": "",
                "Remarks": "SUMMARY",
            })
            continue

        # ── Account block ──
        # Parse primary row
        acct_tok = first_tok
        dates = [(x, v) for x, v in primary_tok_texts if _DATE_PAT.match(v)]
        op_date = dates[0][1] if len(dates) > 0 else ""
        due_date = dates[1][1] if len(dates) > 1 else ""

        large_floats = [(x, v) for x, v in primary_tok_texts
                        if _FLOAT_PAT.match(v) and x > 250 and "." in v]
        sanctioned = large_floats[0][1] if large_floats else ""
        neg_floats = [(x, v) for x, v in large_floats if v.startswith("-")]
        outstanding = neg_floats[0][1] if neg_floats else ""
        pos_after_neg = [(x, v) for x, v in large_floats
                         if not v.startswith("-")
                         and x > (neg_floats[0][0] if neg_floats else 0)]
        overdue_amt = pos_after_neg[0][1] if pos_after_neg else ""

        rates = [(x, v) for x, v in primary_tok_texts
                 if _FLOAT_PAT.match(v) and "." in v and 480 < x < 570
                 and float(v) < 25]
        int_rate = rates[0][1] if rates else ""

        small_ints = [(x, v) for x, v in primary_tok_texts
                      if _INT_PAT.match(v) and x > 720]
        pending_inst = small_ints[0][1] if len(small_ints) > 0 else ""
        overdue_period = small_ints[1][1] if len(small_ints) > 1 else ""

        first_num_x = large_floats[0][0] if large_floats else 999
        name_toks = [v for x, v in primary_tok_texts
                     if x > 60 and x < first_num_x
                     and not _DATE_PAT.match(v) and not _FLOAT_PAT.match(v)]
        name = " ".join(name_toks)

        # Collect all tokens from secondary rows in this block
        sec_tokens_all = []
        for _, _, _, sec_toks in block[1:]:
            sec_tokens_all.extend(sec_toks)

        # Extract fields from merged secondary tokens
        npa_code = next((t[2] for t in sec_tokens_all if _NPA_CODES.match(t[2])), "")
        contact = next((t[2] for t in sec_tokens_all if _CONTACT_PAT.match(t[2])), "")
        floats_sec = [t[2] for t in sec_tokens_all
                      if _FLOAT_PAT.match(t[2]) and t[2] not in (npa_code, contact)]
        cat_toks = [t[2] for t in sec_tokens_all
                    if t[2] not in (npa_code, contact)
                    and not _FLOAT_PAT.match(t[2])
                    and not _INT_PAT.match(t[2])
                    and not _CONTACT_PAT.match(t[2])]

        acct_rows.append({
            "Loan Type Code": ltc, "Loan Type Name": ltn,
            "A/c No.": acct_tok, "Name": name,
            "NPA Code": npa_code, "Category": " ".join(cat_toks),
            "Sanctioned Amt (INR)": sanctioned,
            "Op. Date": op_date, "Due Date": due_date,
            "Int. Rate (%)": int_rate,
            "Outstanding (INR)": outstanding, "Overdue Amt (INR)": overdue_amt,
            "Pending Inst.": pending_inst, "Overdue Period": overdue_period,
            "Adv. Recovery (INR)": floats_sec[0] if len(floats_sec) > 0 else "",
            "Installment Amt (INR)": floats_sec[1] if len(floats_sec) > 1 else "",
            "IR Balance (INR)": "", "Security Amt (INR)": "",
            "Contact No.": contact, "Customer Id": "", "Remarks": "",
        })

    # ── Page 2 ──────────────────────────────────────────────────────────────
    # Page 2: pairs of rows per account
    #   Row A (top row): IR Balance  |  Security Amount
    #   Row B (bottom row): Contact No  |  Customer Id
    # Total rows on page 2 match account count + summary rows

    p2_entries = []  # list of dicts with ir_balance, security_amt, customer_id
    with pdfplumber.open(pdf_path) as pdf:
        if len(pdf.pages) < 2:
            p2_char_rows = []
        else:
            p2_char_rows = _page_rows(pdf.pages[1])

    skip_headers = {"IR", "Balance", "Security", "Amount", "Contact", "No.", "Customer", "Id"}
    p2_pair = {}

    for _, tokens in p2_char_rows:
        toks = [t[2] for t in tokens if t[2] not in skip_headers]
        if not toks:
            continue

        # Row B: contact no (9-10 digits) + customer id (6 digits)
        contacts = [t for t in toks if _CONTACT_PAT.match(t)]
        cust_ids = [t for t in toks if re.match(r"^\d{5,7}$", t) and t not in contacts]
        if contacts or cust_ids:
            p2_pair["contact_p2"] = contacts[0] if contacts else ""
            p2_pair["customer_id"] = cust_ids[0] if cust_ids else ""
            p2_entries.append(dict(p2_pair))
            p2_pair = {}
            continue

        # Row A: two floats (IR balance and security amount)
        floats = [t for t in toks if _FLOAT_PAT.match(t)]
        if floats:
            # Handle merged token bug: "0.0077750000.00"
            real_floats = []
            for f in floats:
                # Detect if two floats merged: has two decimal points
                if f.count(".") == 2:
                    m = re.match(r"^(-?[\d]+\.[\d]+)([\d]+\.[\d]+)$", f)
                    if m:
                        real_floats.extend([m.group(1), m.group(2)])
                        continue
                real_floats.append(f)
            p2_pair["ir_balance"] = real_floats[0] if len(real_floats) > 0 else ""
            p2_pair["security_amt"] = real_floats[1] if len(real_floats) > 1 else ""
            # If this row also has a contact (no customer id row follows), flush immediately
            # (happens for summary rows on page 2)
            if len(real_floats) == 1 and not any(_CONTACT_PAT.match(t) for t in toks):
                p2_entries.append(dict(p2_pair))
                p2_pair = {}

    # ── Merge Page 1 + Page 2 ───────────────────────────────────────────────
    for i, row in enumerate(acct_rows):
        if i < len(p2_entries):
            e = p2_entries[i]
            row["IR Balance (INR)"] = e.get("ir_balance", "")
            row["Security Amt (INR)"] = e.get("security_amt", "")
            row["Customer Id"] = e.get("customer_id", "")
            if not row["Contact No."]:
                row["Contact No."] = e.get("contact_p2", "")

    all_rows = acct_rows + summary_rows
    columns = [
        "Loan Type Code", "Loan Type Name",
        "A/c No.", "Name", "NPA Code", "Category",
        "Sanctioned Amt (INR)", "Op. Date", "Due Date", "Int. Rate (%)",
        "Outstanding (INR)", "Overdue Amt (INR)", "Pending Inst.", "Overdue Period",
        "Adv. Recovery (INR)", "Installment Amt (INR)",
        "IR Balance (INR)", "Security Amt (INR)",
        "Contact No.", "Customer Id", "Remarks",
    ]
    df = pd.DataFrame(all_rows, columns=columns)
    df.attrs["report_date"] = report_date
    return df


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL WRITER
# ─────────────────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
_SUMMARY_FILL = PatternFill("solid", fgColor="FFF2CC")
_SUMMARY_FONT = Font(bold=True, name="Arial", size=9, color="1F4E79")
_DENOM_FILL = PatternFill("solid", fgColor="EBF3FB")
_DATA_FONT = Font(name="Arial", size=9)
_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center")


def _write_sheet(ws, df, title, report_date=""):
    header_text = f"{title}  |  As On: {report_date}" if report_date else title

    # Row 1: Title
    ws.append([header_text])
    cell = ws.cell(row=1, column=1)
    cell.font = Font(bold=True, name="Arial", size=12, color="1F4E79")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    ws.row_dimensions[1].height = 20

    # Row 2: blank
    ws.append([])

    # Row 3: Column headers
    ws.append(list(df.columns))
    ws.row_dimensions[3].height = 28
    for col_idx in range(1, len(df.columns) + 1):
        c = ws.cell(row=3, column=col_idx)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = _CENTER
        c.border = _BORDER

    # Data rows
    for r_idx, row in enumerate(df.itertuples(index=False), start=4):
        is_summary = str(getattr(row, "Remarks", "")).strip() in ("SUMMARY", "SUMMARY ROW")
        is_section_label = any(
            kw in str(row[2]) for kw in
            ["Type Wise", "Branch Wise", "Grand Total", "Typewise", "Branchwise", "TOTAL"]
        )
        is_denom = str(row[0]) == "Denomination Breakup" if df.columns[0] == "Section" else False

        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=str(value) if value else "")
            cell.border = _BORDER
            if is_summary or is_section_label:
                cell.fill = _SUMMARY_FILL
                cell.font = _SUMMARY_FONT
                cell.alignment = _CENTER
            elif is_denom:
                cell.fill = _DENOM_FILL
                cell.font = _DATA_FONT
                cell.alignment = _LEFT
            else:
                cell.font = _DATA_FONT
                cell.alignment = _LEFT

    # Auto column widths
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_vals = df.iloc[:, col_idx - 1].astype(str)
        max_len = max(len(str(col_name)), col_vals.map(len).max() if not df.empty else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 35)

    ws.freeze_panes = "A4"


def generate_master_excel(cash_df, insurance_df, overdue_df, output_path):
    wb = Workbook()
    wb.remove(wb.active)

    sheets = [
        ("Cash Summary",     cash_df,      "Cash Summary Report – PANCHOT BRANCH"),
        ("Insurance Pending", insurance_df, "Insurance Pending Register – PANCHOT BRANCH"),
        ("Overdue Report",   overdue_df,   "Overdue Report For All – PANCHOT BRANCH"),
    ]

    for sheet_name, df, title in sheets:
        ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, df, title, df.attrs.get("report_date", ""))

    wb.save(output_path)
    print(f"Saved → {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    UPLOAD_DIR = "/mnt/user-data/uploads"
    OUTPUT_DIR = "outputs"
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    FILES = {
        "cash":      "D:\\audit_report_generator\\uploads\\PANCHOT BR CASH FEB SUM.pdf",
        "insurance": "D:\\audit_report_generator\\uploads\\PANCHOT BR JAN 2026 INSURANCE PENDING REGI.pdf",
        "overdue":   "D:\\audit_report_generator\\uploads\\PANCHOT BR JAN 2026 OVER DUE REPORT.pdf",
    }

    print("Extracting Cash Summary...")
    cash_df = extract_cash_summary(os.path.join(UPLOAD_DIR, FILES["cash"]))
    print(f"  → {len(cash_df)} rows extracted")

    print("Extracting Insurance Pending Register...")
    ins_df = extract_insurance_pending(os.path.join(UPLOAD_DIR, FILES["insurance"]))
    print(f"  → {len(ins_df)} rows extracted")

    print("Extracting Overdue Report...")
    ov_df = extract_overdue_report(os.path.join(UPLOAD_DIR, FILES["overdue"]))
    print(f"  → {len(ov_df)} rows extracted")

    output_path = os.path.join(OUTPUT_DIR, "PANCHOT_MASTER_REPORT.xlsx")
    print(f"\nGenerating Excel workbook...")
    generate_master_excel(cash_df, ins_df, ov_df, output_path)

    print("\n── Cash Summary ──────────────────────────────")
    print(cash_df.to_string(index=False))
    print("\n── Insurance Pending ─────────────────────────")
    print(ins_df.to_string(index=False))
    print("\n── Overdue Report ────────────────────────────")
    print(ov_df.to_string(index=False))
