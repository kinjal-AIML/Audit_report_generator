import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict
from pathlib import Path
from utils.logger import logger

class ExcelReconstructionEngine:
    """
    Module 9: Excel Reconstruction Engine
    Generates MASTER_AUDIT_DATA.xlsx with dynamically resized columns and unified relational mapping keys.
    """
    def __init__(self, output_dir: str = "extracted/excel"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.header_fill = PatternFill("solid", fgColor="1F4E79")
        self.header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        self.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        self.summary_fill = PatternFill("solid", fgColor="FFF2CC")
        self.summary_font = Font(bold=True, name="Arial", size=9, color="1F4E79")
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.left_align = Alignment(horizontal="left", vertical="center")
        
    def generate_master_excel(self, tables: Dict[str, pd.DataFrame], filename: str = "MASTER_AUDIT_DATA.xlsx") -> str:
        logger.info(f"Generating master Excel: {filename}")
        wb = Workbook()
        wb.remove(wb.active)
        
        for sheet_name, df in tables.items():
            if df.empty:
                logger.warning(f"Sheet {sheet_name} is empty, skipping.")
                continue
            
            # Ensure relational mapping columns exist
            for key in ["Remarks", "Annexure Ref", "Observation ID", "Compliance Status"]:
                if key not in df.columns:
                    df[key] = ""
                    
            ws = wb.create_sheet(title=sheet_name[:31]) # Excel sheet name limit
            
            # Write Headers
            ws.append(list(df.columns))
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.border = self.border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            # Write Data
            for row in df.itertuples(index=False):
                ws.append([str(v) if pd.notnull(v) else "" for v in row])
                
            # Formatting
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                # Using Remarks column (which is close to the end, or typically contains SUMMARY string)
                # Fallback to checking the first few columns
                cell_values = [str(c.value).upper() for c in row if c.value]
                is_summary = any("SUMMARY" in v or "TOTAL" in v for v in cell_values)
                
                for cell in row:
                    cell.border = self.border
                    if is_summary:
                        cell.fill = self.summary_fill
                        cell.font = self.summary_font
                        cell.alignment = self.center_align
                    else:
                        cell.alignment = self.left_align
                    
            # Auto-size columns
            for col_idx, col_name in enumerate(df.columns, start=1):
                max_len = max([len(str(v)) for v in df.iloc[:, col_idx - 1].dropna()] + [len(str(col_name))])
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)
        
        output_path = self.output_dir / filename
        wb.save(str(output_path))
        logger.info(f"Excel File Saved: {output_path}")
        return str(output_path)
