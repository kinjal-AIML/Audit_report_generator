from typing import Union
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from docx import Document
from docx.oxml.text.paragraph import CT_P
from docx.oxml.table import CT_Tbl
from docx.table import _Cell, Table
from docx.text.paragraph import Paragraph


def _iter_block_items(parent):
    if isinstance(parent, _Cell):
        parent_elm = parent._tc
    else:
        parent_elm = parent.element.body
    for child in parent_elm.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, parent)
        elif isinstance(child, CT_Tbl):
            yield Table(child, parent)


def _first_run_font(paragraph: Paragraph):
    name = None
    size_pt = None
    bold = False
    italic = False
    for r in paragraph.runs:
        if r.font is None:
            continue
        if name is None and r.font.name:
            name = r.font.name
        if size_pt is None and r.font.size:
            try:
                size_pt = r.font.size.pt
            except Exception:
                size_pt = None
        bold = bold or bool(r.bold)
        italic = italic or bool(r.italic)
    # Fallback to paragraph style
    if paragraph.style and paragraph.style.font:
        if name is None and paragraph.style.font.name:
            name = paragraph.style.font.name
        if size_pt is None and paragraph.style.font.size:
            try:
                size_pt = paragraph.style.font.size.pt
            except Exception:
                size_pt = None
    return name or "Calibri", size_pt or 11, bold, italic


def generate_docx_like_excel(docx_path: str, output_dir: str) -> str:
    """
    Render the DOCX layout into a single Excel sheet, preserving reading order
    and approximating basic formatting (font name, size, bold/italic). Paragraphs
    are written as merged, wrapped rows; tables are written as tabular blocks.
    """
    doc = Document(docx_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "DOCX"

    # Wider sheet, wrapped text by default
    for col in range(1, 9):  # A..H
        ws.column_dimensions[chr(64 + col)].width = 28

    wrap = Alignment(wrap_text=True, vertical='top', horizontal='left')

    row = 1
    for block in _iter_block_items(doc):
        if isinstance(block, Paragraph):
            text = block.text.strip()
            # skip empty paragraphs but keep minimal spacing for visual parity
            if not text:
                row += 1
                continue
            font_name, size_pt, bold, italic = _first_run_font(block)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            cell = ws.cell(row=row, column=1, value=text)
            cell.font = Font(name=font_name, size=size_pt, bold=bold or block.style.name.startswith("Heading") if block.style else bold, italic=italic)
            cell.alignment = wrap
            row += 2  # spacing after paragraph
        else:  # Table
            table: Table = block
            n_cols = max((len(r.cells) for r in table.rows), default=1)
            # Write rows
            for r_idx, r in enumerate(table.rows):
                for c_idx, c in enumerate(r.cells):
                    text = c.text.strip()
                    cell = ws.cell(row=row + r_idx, column=1 + c_idx, value=text)
                    # Header style for first row
                    if r_idx == 0:
                        cell.font = Font(bold=True)
                    cell.alignment = wrap
            row += len(table.rows) + 2  # spacing after table

    os.makedirs(output_dir, exist_ok=True)
    out_path = os.path.join(output_dir, "DOCX_CONTENT.xlsx")
    wb.save(out_path)
    return out_path
