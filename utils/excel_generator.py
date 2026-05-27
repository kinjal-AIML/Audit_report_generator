import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL STYLING CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
_SUMMARY_FILL = PatternFill("solid", fgColor="FFF2CC")
_SUMMARY_FONT = Font(bold=True, name="Arial", size=9, color="1F4E79")
_TOTAL_FILL   = PatternFill("solid", fgColor="D6E4F0")
_TOTAL_FONT   = Font(bold=True, name="Arial", size=9, color="1F4E79")
_SECTION_FILL = PatternFill("solid", fgColor="EBF3FB")
_DATA_FONT    = Font(name="Arial", size=9)
_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center")

def _write_sheet(ws, df, title, report_date=""):
    """
    Writes a styled sheet to the workbook from a DataFrame.
    """
    header_text = f"{title}  |  Period: {report_date}" if report_date else title

    # Main Header
    ws.append([header_text])
    cell = ws.cell(row=1, column=1)
    cell.font = Font(bold=True, name="Arial", size=12, color="1F4E79")
    num_cols = len(df.columns) if len(df.columns) > 0 else 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    ws.row_dimensions[1].height = 20

    ws.append([]) # Spacer row

    # Column Headers
    ws.append(list(df.columns))
    ws.row_dimensions[3].height = 28
    for col_idx in range(1, len(df.columns) + 1):
        c = ws.cell(row=3, column=col_idx)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = _CENTER
        c.border = _BORDER

    # Data Rows
    for r_idx, row in enumerate(df.itertuples(index=False), start=4):
        # Determine row style based on content or 'Remarks'
        row_str = " ".join(str(v) for v in row).upper()
        remarks = str(getattr(row, "Remarks", "")).strip().upper()
        
        is_summary = remarks in ("SUMMARY", "SUMMARY ROW")
        is_total = "****TOTAL****" in row_str or "GRAND TOTAL" in row_str
        is_section_label = any(kw in row_str for kw in ["TYPE WISE", "BRANCH WISE", "TOTAL FOR"])
        
        # Specific Trial Balance check
        side_val = str(getattr(row, "Side", "")).strip() if hasattr(row, "Side") else ""
        is_side_header = side_val in ("LIABILITIES", "ASSETS", "INCOME", "EXPENDITURE", "NET PROFIT")

        for c_idx, value in enumerate(row, start=1):
            display = "" if str(value) in ("False", "True", "nan") else str(value)
            cell = ws.cell(row=r_idx, column=c_idx, value=display)
            cell.border = _BORDER

            if is_total:
                cell.fill = _TOTAL_FILL
                cell.font = _TOTAL_FONT
                cell.alignment = _CENTER
            elif is_summary or is_section_label:
                cell.fill = _SUMMARY_FILL
                cell.font = _SUMMARY_FONT
                cell.alignment = _CENTER
            elif is_side_header:
                cell.fill = _SECTION_FILL
                cell.font = Font(bold=True, name="Arial", size=9, color="1F4E79")
                cell.alignment = _LEFT
            else:
                cell.font = _DATA_FONT
                cell.alignment = _LEFT

    # Column Widths
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_vals = df.iloc[:, col_idx - 1].astype(str)
        max_len = max(len(str(col_name)), col_vals.map(len).max() if not df.empty else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A4"

def generate_master_excel(data_dict, output_dir):
    """
    Creates MASTER_AUDIT_DATA.xlsx with multiple styled sheets.
    data_dict: { 'Sheet Name': df, ... }
    """
    output_path = os.path.join(output_dir, "MASTER_AUDIT_DATA.xlsx")
    wb = Workbook()
    wb.remove(wb.active) # Remove default sheet

    for sheet_name, df in data_dict.items():
        # Ensure required columns exist for the general app logic
        if "Remarks" not in df.columns:
            df["Remarks"] = ""
        if "Annexure Ref" not in df.columns:
            df["Annexure Ref"] = ""
            
        ws = wb.create_sheet(title=sheet_name[:31])
        report_date = df.attrs.get("report_date", "")
        _write_sheet(ws, df, sheet_name, report_date)

    wb.save(output_path)
    return output_path

def generate_annexure_workbook(data_dict, output_dir):
    """
    Creates separate Annexure Workbook if needed.
    """
    output_path = os.path.join(output_dir, "ANNEXURE_WORKBOOK.xlsx")
    wb = Workbook()
    wb.remove(wb.active)
    
    df_annexures = data_dict.get("Annexures", pd.DataFrame())
    if not df_annexures.empty:
        ws = wb.create_sheet(title="Annexures")
        _write_sheet(ws, df_annexures, "Annexures")
    else:
        # Create an empty sheet if no data
        wb.create_sheet(title="Annexures")
        
    wb.save(output_path)
    return output_path
