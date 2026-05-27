"""
PANCHOT Branch Report Extractor → Excel Master Sheet
=====================================================
Extracts structured data from 5 bank branch PDFs into a formatted
multi-sheet Excel workbook.

Sheets produced:
  1. Cash Summary          ← PANCHOT_BR_CASH_FEB_SUM.pdf
  2. Insurance Pending     ← PANCHOT_BR_JAN_2026_INSURANCE_PENDING_REGI_.pdf
  3. Overdue Report        ← PANCHOT_BR_JAN_2026_OVER_DUE_REPORT.pdf
  4. Limit Register        ← PANCHOT_BR_JAN_2026_LIMIT_REGI_.pdf
  5. Trial Balance         ← PANCHOT_BR_JAN_2026_TRIAL_BALANCE.pdf

Requirements:
    pip install pdfplumber openpyxl

Usage:
    python panchot_report_extractor.py
"""

import os
import re
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# CHAR-LEVEL WORD RECONSTRUCTION
# PDFs use spaced-character font encoding: individual letters have 0.01–0.11 px
# gaps within a visual word and ~5.5 px gaps between visual words.
# Standard pdfplumber text extraction cannot handle this correctly.
# ─────────────────────────────────────────────────────────────────────────────

INTRA_WORD_GAP_PX = 3.0    # gap within same visual word token
INTER_WORD_GAP_PX = 15.0   # gap above which = column/field boundary


def _chars_to_words(chars):
    """
    Group page chars into word tokens using pixel-gap thresholds.
    Returns: list of (x0, x1, text) tuples.
    """
    if not chars:
        return []
    chars_sorted = sorted(chars, key=lambda c: c["x0"])
    tokens = []
    buf = [chars_sorted[0]]
    for c in chars_sorted[1:]:
        gap = c["x0"] - buf[-1]["x1"]
        if gap <= INTRA_WORD_GAP_PX:
            buf.append(c)
        else:
            tokens.append((buf[0]["x0"], buf[-1]["x1"], "".join(ch["text"] for ch in buf)))
            buf = [c]
    tokens.append((buf[0]["x0"], buf[-1]["x1"], "".join(ch["text"] for ch in buf)))
    return tokens


def _page_rows(page, y_group_tolerance=3):
    """
    Return ordered list of (top, tokens) rows, grouped by y-position.
    """
    chars = [c for c in page.chars if c["text"].strip()]
    row_map = {}
    for c in chars:
        key = round(c["top"] / y_group_tolerance) * y_group_tolerance
        row_map.setdefault(key, []).append(c)
    rows = []
    for top in sorted(row_map):
        tokens = _chars_to_words(row_map[top])
        if tokens:
            rows.append((top, tokens))
    return rows


def _all_rows_multipage(pdf_path):
    """Return (page_idx, top, tokens) for all pages."""
    result = []
    with pdfplumber.open(pdf_path) as pdf:
        for pi, page in enumerate(pdf.pages):
            for top, tokens in _page_rows(page):
                result.append((pi, top, tokens))
    return result


def _token_text(tokens):
    return " ".join(t[2] for t in tokens)


# ─────────────────────────────────────────────────────────────────────────────
# NOISE FILTER
# ─────────────────────────────────────────────────────────────────────────────

_NOISE = re.compile(
    r"THE MEHSANA|PANCHOT BRANCH|MICR:-|IFSC:-|Print Date|"
    r"^-{5,}$|User Name:|Page \d+ of|^Clerk\s|^Cashier\s|"
    r"Security Details|A/c No\.\s+Name\s+NPA|"
    r"Overdue Report For All|Cash Summary Report|Insurance Pending Register|"
    r"Limit Register From|General Ledger/Trial|Profit & loss ledger|"
    r"^A/c\.\s+Holder|^Security\s*$|^Ac/No\s+Head|"
    r"GL TRIAL BALANCE TALLY",
    re.IGNORECASE,
)


def _is_noise(text):
    return bool(_NOISE.search(text))


# ─────────────────────────────────────────────────────────────────────────────
# COMMON PATTERNS
# ─────────────────────────────────────────────────────────────────────────────

_DATE_PAT  = re.compile(r"^\d{2}/\d{2}/\d{4}$")
_FLOAT_PAT = re.compile(r"^-?[\d,]+\.?\d*$")
_INT_PAT   = re.compile(r"^\d+$")
_CONTACT_PAT = re.compile(r"^\d{9,10}$")
_NPA_CODES = re.compile(r"^(SMA\d|0[0-9]|\d{2}|NPA)$")
_ACCT6     = re.compile(r"^\d{6}$")
_AMOUNT    = re.compile(r"^-?[\d,]+\.\d{2}$")


# ─────────────────────────────────────────────────────────────────────────────
# 1. CASH SUMMARY
# ─────────────────────────────────────────────────────────────────────────────

def extract_cash_summary(pdf_path):
    rows = []
    report_date = ""

    for _, _, tokens in _all_rows_multipage(pdf_path):
        line = _token_text(tokens)
        if _is_noise(line):
            continue

        m = re.search(r"AsOn\s+(\S+)", line, re.IGNORECASE)
        if m:
            report_date = m.group(1)
            continue

        m = re.match(
            r"^(Opening Cash Balance|Total Cash Receipt|Total Cash Payment|Closing Cash Balance)\s+([\d,\.]+)$",
            line,
        )
        if m:
            rows.append({"Section": "Cash Flow", "Description": m.group(1),
                         "Count": "", "Amount (INR)": m.group(2)})
            continue

        if len(tokens) >= 3:
            last = tokens[-1][2]
            second_last = tokens[-2][2]
            if _AMOUNT.match(last) and _INT_PAT.match(second_last):
                label = " ".join(t[2] for t in tokens[:-2])
                if any(kw in label.upper() for kw in ["NOTE", "COIN", "PAISA"]):
                    rows.append({"Section": "Denomination Breakup", "Description": label,
                                 "Count": second_last, "Amount (INR)": last})
                    continue

        m = re.match(r"^Total:\s*([\d,\.]+)$", line)
        if m:
            rows.append({"Section": "Denomination Breakup", "Description": "TOTAL",
                         "Count": "", "Amount (INR)": m.group(1)})
            continue

        m = re.search(r"Branch Cash Retention Limit\s*:\s*([\d,\.]+)", line)
        if m:
            rows.append({"Section": "Info", "Description": "Branch Cash Retention Limit",
                         "Count": "", "Amount (INR)": m.group(1)})

    df = pd.DataFrame(rows, columns=["Section", "Description", "Count", "Amount (INR)"])
    df.attrs["report_date"] = report_date
    return df


# ─────────────────────────────────────────────────────────────────────────────
# 2. INSURANCE PENDING REGISTER
# ─────────────────────────────────────────────────────────────────────────────

def extract_insurance_pending(pdf_path):
    rows = []
    report_date = ""
    branch_code = branch_name = loan_type_code = loan_type_name = ""

    for _, _, tokens in _all_rows_multipage(pdf_path):
        line = _token_text(tokens)
        if not line.strip() or _is_noise(line):
            continue

        m = re.search(r"As On\s*:\s*(\S+)", line, re.IGNORECASE)
        if m:
            report_date = m.group(1)
            continue

        m = re.match(r"^\((\d+)\)\s*-\s*(.+)$", line)
        if m:
            code, name = m.group(1).strip(), m.group(2).strip()
            if code == "004":
                branch_code, branch_name = code, name
            else:
                loan_type_code, loan_type_name = code, name
            continue

        m = re.search(
            r"(Type Wise Total|Branch Wise Total|Grand Total).+?:\s*(\d+)\s+([\d,\.]+)",
            line, re.IGNORECASE,
        )
        if m:
            rows.append({
                "Branch Code": branch_code, "Branch Name": branch_name,
                "Loan Type Code": "", "Loan Type Name": "",
                "A/c No.": "", "Name": m.group(1).strip(),
                "NPA Code": "", "Limit Amount (INR)": m.group(3),
                "Remarks": f"Count: {m.group(2)}",
            })
            continue

        if not tokens:
            continue
        acct_tok = tokens[0][2]
        if not _ACCT6.match(acct_tok):
            continue
        amt_tok = tokens[-1][2]
        if not _AMOUNT.match(amt_tok):
            continue

        npa_code = ""
        name_tokens = []
        for i in range(len(tokens) - 2, 0, -1):
            if _NPA_CODES.match(tokens[i][2]):
                npa_code = tokens[i][2]
                name_tokens = tokens[1:i]
                break
        else:
            name_tokens = tokens[1:-1]

        rows.append({
            "Branch Code": branch_code, "Branch Name": branch_name,
            "Loan Type Code": loan_type_code, "Loan Type Name": loan_type_name,
            "A/c No.": acct_tok,
            "Name": " ".join(t[2] for t in name_tokens),
            "NPA Code": npa_code, "Limit Amount (INR)": amt_tok,
            "Remarks": "",
        })

    df = pd.DataFrame(rows, columns=[
        "Branch Code", "Branch Name", "Loan Type Code", "Loan Type Name",
        "A/c No.", "Name", "NPA Code", "Limit Amount (INR)", "Remarks",
    ])
    df.attrs["report_date"] = report_date
    return df


# ─────────────────────────────────────────────────────────────────────────────
# 3. OVERDUE REPORT
# ─────────────────────────────────────────────────────────────────────────────

_SUMMARY_PAT = re.compile(r"^(Typewise|Branchwise|Grand)", re.IGNORECASE)


def extract_overdue_report(pdf_path):
    report_date = ""
    acct_rows = []
    summary_rows = []
    loan_type_code = loan_type_name = ""

    with pdfplumber.open(pdf_path) as pdf:
        p1_char_rows = _page_rows(pdf.pages[0])

    # ── Segment rows into per-account blocks ─────────────────────────────────
    blocks = []
    current_block = []

    for top, tokens in p1_char_rows:
        line = _token_text(tokens)
        if not line.strip():
            continue

        m = re.search(r"As On\s+(\d{2}/\d{2}/\d{4})", line, re.IGNORECASE)
        if m and not report_date:
            report_date = m.group(1)
            continue

        if _is_noise(line):
            continue

        m = re.match(r"^\((\d+\s*)\)\s*-\s*(.+)$", line)
        if m:
            code, name = m.group(1).strip(), m.group(2).strip()
            if code != "004":
                loan_type_code, loan_type_name = code, name
            continue

        first_tok = tokens[0][2] if tokens else ""
        is_new_acct = _ACCT6.match(first_tok)
        is_summary  = _SUMMARY_PAT.match(first_tok)

        if is_new_acct or is_summary:
            if current_block:
                blocks.append(current_block)
            current_block = [(loan_type_code, loan_type_name, top, tokens)]
        else:
            if current_block:
                current_block.append((loan_type_code, loan_type_name, top, tokens))

    if current_block:
        blocks.append(current_block)

    # ── Parse each block ──────────────────────────────────────────────────────
    for block in blocks:
        if not block:
            continue

        ltc, ltn, _, primary_tokens = block[0]
        primary_tok_texts = [(t[0], t[2]) for t in primary_tokens]
        first_tok = primary_tokens[0][2] if primary_tokens else ""

        if _SUMMARY_PAT.match(first_tok):
            line = _token_text(primary_tokens)
            nums = [v for _, v in primary_tok_texts if _FLOAT_PAT.match(v)]
            ints = [v for _, v in primary_tok_texts if _INT_PAT.match(v) and v not in nums]
            label = re.match(r"^([\w\s]+Total)", line, re.IGNORECASE)
            summary_rows.append({
                "Loan Type Code": "", "Loan Type Name": "",
                "A/c No.": label.group(1).strip() if label else "Total",
                "Name": f"({ints[0]} accounts)" if ints else "",
                "NPA Code": "", "Category": "",
                "Sanctioned Amt (INR)": nums[0] if len(nums) > 0 else "",
                "Op. Date": "", "Due Date": "", "Int. Rate (%)": "",
                "Adv. Recovery (INR)": nums[1] if len(nums) > 1 else "",
                "Installment Amt (INR)": nums[2] if len(nums) > 2 else "",
                "Outstanding (INR)": nums[3] if len(nums) > 3 else "",
                "Overdue Amt (INR)": nums[4] if len(nums) > 4 else "",
                "Pending Inst.": "", "Overdue Period": "",
                "IR Balance (INR)": nums[5] if len(nums) > 5 else "",
                "Security Amt (INR)": "", "Contact No.": "", "Customer Id": "",
                "Remarks": "SUMMARY",
            })
            continue

        acct_tok = first_tok
        dates = [(x, v) for x, v in primary_tok_texts if _DATE_PAT.match(v)]
        op_date  = dates[0][1] if len(dates) > 0 else ""
        due_date = dates[1][1] if len(dates) > 1 else ""

        large_floats = [(x, v) for x, v in primary_tok_texts
                        if _FLOAT_PAT.match(v) and x > 250 and "." in v]
        sanctioned  = large_floats[0][1] if large_floats else ""
        neg_floats  = [(x, v) for x, v in large_floats if v.startswith("-")]
        outstanding = neg_floats[0][1] if neg_floats else ""
        pos_after   = [(x, v) for x, v in large_floats
                       if not v.startswith("-") and x > (neg_floats[0][0] if neg_floats else 0)]
        overdue_amt = pos_after[0][1] if pos_after else ""

        rates = [(x, v) for x, v in primary_tok_texts
                 if _FLOAT_PAT.match(v) and "." in v and 480 < x < 570 and float(v) < 25]
        int_rate = rates[0][1] if rates else ""

        small_ints = [(x, v) for x, v in primary_tok_texts if _INT_PAT.match(v) and x > 720]
        pending_inst  = small_ints[0][1] if len(small_ints) > 0 else ""
        overdue_period = small_ints[1][1] if len(small_ints) > 1 else ""

        first_num_x = large_floats[0][0] if large_floats else 999
        name_toks = [v for x, v in primary_tok_texts
                     if x > 60 and x < first_num_x
                     and not _DATE_PAT.match(v) and not _FLOAT_PAT.match(v)]
        name = " ".join(name_toks)

        sec_tokens_all = []
        for _, _, _, sec_toks in block[1:]:
            sec_tokens_all.extend(sec_toks)

        npa_code = next((t[2] for t in sec_tokens_all if _NPA_CODES.match(t[2])), "")
        contact  = next((t[2] for t in sec_tokens_all if _CONTACT_PAT.match(t[2])), "")
        floats_sec = [t[2] for t in sec_tokens_all
                      if _FLOAT_PAT.match(t[2]) and t[2] not in (npa_code, contact)]
        cat_toks = [t[2] for t in sec_tokens_all
                    if t[2] not in (npa_code, contact)
                    and not _FLOAT_PAT.match(t[2]) and not _INT_PAT.match(t[2])
                    and not _CONTACT_PAT.match(t[2])]

        acct_rows.append({
            "Loan Type Code": ltc, "Loan Type Name": ltn,
            "A/c No.": acct_tok, "Name": name,
            "NPA Code": npa_code, "Category": " ".join(cat_toks),
            "Sanctioned Amt (INR)": sanctioned,
            "Op. Date": op_date, "Due Date": due_date,
            "Int. Rate (%)": int_rate,
            "Outstanding (INR)": outstanding, "Overdue Amt (INR)": overdue_amt,
            "Pending Inst.": pending_inst, "Overdue Period": overdue_period,
            "Adv. Recovery (INR)": floats_sec[0] if len(floats_sec) > 0 else "",
            "Installment Amt (INR)": floats_sec[1] if len(floats_sec) > 1 else "",
            "IR Balance (INR)": "", "Security Amt (INR)": "",
            "Contact No.": contact, "Customer Id": "", "Remarks": "",
        })

    # ── Page 2: IR Balance, Security Amt, Customer Id ────────────────────────
    p2_entries = []
    with pdfplumber.open(pdf_path) as pdf:
        p2_char_rows = _page_rows(pdf.pages[1]) if len(pdf.pages) > 1 else []

    skip_hdrs = {"IR", "Balance", "Security", "Amount", "Contact", "No.", "Customer", "Id"}
    p2_pair = {}

    for _, tokens in p2_char_rows:
        toks = [t[2] for t in tokens if t[2] not in skip_hdrs]
        if not toks:
            continue

        contacts = [t for t in toks if _CONTACT_PAT.match(t)]
        cust_ids = [t for t in toks if re.match(r"^\d{5,7}$", t) and t not in contacts]
        if contacts or cust_ids:
            p2_pair["contact_p2"] = contacts[0] if contacts else ""
            p2_pair["customer_id"] = cust_ids[0] if cust_ids else ""
            p2_entries.append(dict(p2_pair))
            p2_pair = {}
            continue

        floats = [t for t in toks if _FLOAT_PAT.match(t)]
        if floats:
            real_floats = []
            for f in floats:
                if f.count(".") == 2:
                    m = re.match(r"^(-?[\d]+\.[\d]+)([\d]+\.[\d]+)$", f)
                    if m:
                        real_floats.extend([m.group(1), m.group(2)])
                        continue
                real_floats.append(f)
            p2_pair["ir_balance"]   = real_floats[0] if len(real_floats) > 0 else ""
            p2_pair["security_amt"] = real_floats[1] if len(real_floats) > 1 else ""
            if len(real_floats) == 1 and not any(_CONTACT_PAT.match(t) for t in toks):
                p2_entries.append(dict(p2_pair))
                p2_pair = {}

    for i, row in enumerate(acct_rows):
        if i < len(p2_entries):
            e = p2_entries[i]
            row["IR Balance (INR)"]  = e.get("ir_balance", "")
            row["Security Amt (INR)"] = e.get("security_amt", "")
            row["Customer Id"]        = e.get("customer_id", "")
            if not row["Contact No."]:
                row["Contact No."] = e.get("contact_p2", "")

    columns = [
        "Loan Type Code", "Loan Type Name",
        "A/c No.", "Name", "NPA Code", "Category",
        "Sanctioned Amt (INR)", "Op. Date", "Due Date", "Int. Rate (%)",
        "Outstanding (INR)", "Overdue Amt (INR)", "Pending Inst.", "Overdue Period",
        "Adv. Recovery (INR)", "Installment Amt (INR)",
        "IR Balance (INR)", "Security Amt (INR)",
        "Contact No.", "Customer Id", "Remarks",
    ]
    df = pd.DataFrame(acct_rows + summary_rows, columns=columns)
    df.attrs["report_date"] = report_date
    return df


# ─────────────────────────────────────────────────────────────────────────────
# 4. LIMIT REGISTER
# Structure per account (2 rows each):
#   Row A: 000004 TANISH DEVELOPERS  OPEN  NORMAL  01/05/2025  30/04/2026  1500000.00
#   Row B: CASH CREDIT (limit_type)  12.00 (int_rate)  3.00 (over_rate)  [entered_by]  [verified_by]
# ─────────────────────────────────────────────────────────────────────────────

def extract_limit_register(pdf_path):
    rows = []
    report_date_from = report_date_to = ""
    loan_type_code = loan_type_name = ""

    # Known column-header words that appear in repeat page headers — must not be
    # treated as account data. These appear in the secondary (row B) slot when a
    # page break falls between primary and secondary rows of an account.
    _LIMIT_HEADER_WORDS = {
        "Security", "Int.Rate", "Over", "Rate", "Value", "AmountEntered",
        "By", "Verified", "Status", "Limit", "Type", "Sanction",
        "Date", "Due", "Amount", "Holder.", "A/c.",
        "NORMAL",  # appears in header and in data — skip from secondary
    }

    blocks = []
    current_block = []
    summary_rows_local = []

    all_rows = _all_rows_multipage(pdf_path)

    for _, _, tokens in all_rows:
        line = _token_text(tokens)
        if not line.strip() or _is_noise(line):
            continue

        # Date range
        m = re.search(r"From\s*:\s*(\S+)\s+To\s*:\s*(\S+)", line, re.IGNORECASE)
        if m:
            report_date_from = m.group(1)
            report_date_to   = m.group(2)
            continue

        if re.match(r"^PANCHOT BRANCH$", line, re.IGNORECASE):
            continue

        # Skip pure column-header rows (e.g. "Security Int.Rate Over Rate Value AmountEntered By Verified By")
        non_header_toks = [t for t in tokens if t[2] not in _LIMIT_HEADER_WORDS
                           and not _ACCT6.match(t[2]) and not _DATE_PAT.match(t[2])
                           and not _AMOUNT.match(t[2])]
        if all(t[2] in _LIMIT_HEADER_WORDS for t in tokens):
            continue

        m = re.match(r"^Type\s*:\s*\((\d+)\)\s*-\s*(.+)$", line)
        if m:
            if current_block:
                blocks.append(current_block)
            current_block = []
            loan_type_code = m.group(1).strip()
            loan_type_name = m.group(2).strip()
            continue

        # Summary rows
        m = re.match(r"^Total\s+for\s+(Type|Branch)\s*:", line, re.IGNORECASE)
        if m:
            if current_block:
                blocks.append(current_block)
                current_block = []
            nums = [t[2] for t in tokens if _AMOUNT.match(t[2])]
            cnt_toks = [t[2] for t in tokens if _INT_PAT.match(t[2]) and int(t[2]) < 10000]
            summary_rows_local.append({
                "Loan Type Code": loan_type_code,
                "Loan Type Name": loan_type_name,
                "A/c No.": "",
                "Name": re.sub(r"\s+", " ", line[:60]).strip(),
                "Status": "",
                "Sanction Date": "",
                "Due Date": "",
                "Limit Type": "",
                "Int. Rate (%)": "",
                "Over Rate (%)": "",
                "Limit Amount (INR)": nums[-1] if nums else "",
                "Security Value Amt (INR)": nums[-2] if len(nums) > 1 else "",
                "Entered By": "",
                "Verified By": "",
                "Remarks": "SUMMARY",
            })
            continue

        first_tok = tokens[0][2] if tokens else ""
        if _ACCT6.match(first_tok):
            if current_block:
                blocks.append(current_block)
            current_block = [(loan_type_code, loan_type_name, tokens)]
        else:
            # Filter out pure column-header repeat rows before adding to block
            if current_block and not all(t[2] in _LIMIT_HEADER_WORDS for t in tokens):
                current_block.append((loan_type_code, loan_type_name, tokens))

    if current_block:
        blocks.append(current_block)

    # ── Parse blocks ──────────────────────────────────────────────────────────
    for block in blocks:
        if not block:
            continue

        ltc, ltn, primary_tokens = block[0]
        primary_tok_texts = [(t[0], t[2]) for t in primary_tokens]
        first_tok = primary_tokens[0][2]

        if not _ACCT6.match(first_tok):
            continue

        acct_no = first_tok

        # Dates from primary row
        dates = [(x, v) for x, v in primary_tok_texts if _DATE_PAT.match(v)]
        sanction_date = dates[0][1] if len(dates) > 0 else ""
        due_date      = dates[1][1] if len(dates) > 1 else ""

        # Status: "OPEN" or "CLOSE" — x around 263
        status_toks = [v for x, v in primary_tok_texts
                       if v in ("OPEN", "CLOSE", "CLOSED") and 240 < x < 290]
        status = status_toks[0] if status_toks else ""

        # Limit amount: last amount token on primary row
        amounts = [(x, v) for x, v in primary_tok_texts if _AMOUNT.match(v)]
        limit_amount = amounts[-1][1] if amounts else ""

        # Name: tokens between acct_no and STATUS field (x < ~240)
        first_status_x = min((x for x, v in primary_tok_texts
                               if v in ("OPEN", "CLOSE", "CLOSED", "NORMAL")), default=240)
        name_toks = [v for x, v in primary_tok_texts
                     if x > 40 and x < first_status_x
                     and not _DATE_PAT.match(v) and not _AMOUNT.match(v)]
        name = " ".join(name_toks)

        # Collect all secondary rows in this block
        sec_all_tokens = []
        for _, _, sec_toks in block[1:]:
            sec_all_tokens.extend(sec_toks)

        sec_tok_texts = [(t[0], t[2]) for t in sec_all_tokens]

        # Limit type: text tokens on secondary row at x < ~280, not a number, not header words
        limit_type_toks = [v for x, v in sec_tok_texts
                           if x < 280 and not _AMOUNT.match(v)
                           and not _INT_PAT.match(v) and not _FLOAT_PAT.match(v)
                           and v not in _LIMIT_HEADER_WORDS]
        clean_limit_type = " ".join(limit_type_toks)
        # Normalise PDF artefacts in limit type strings
        # "CRASH CREDITCASH CREDIT" → "CASH CREDIT" (print glitch)
        clean_limit_type = re.sub(r"CRASH\s+CREDIT\w*", "", clean_limit_type)
        # "NOCCASH CREDIT" → "CASH CREDIT" (NOC prefix merged with next word)
        clean_limit_type = re.sub(r"\bNOC(\w)", r"\1", clean_limit_type)
        # Strip security description prefixes e.g. "FLAT NO A 302" or "FLAT NO A"
        clean_limit_type = re.sub(r"\bFLAT\s+NO\s+\S+\s*", "", clean_limit_type)
        clean_limit_type = re.sub(r"HYPOHYPO", "HYPO", clean_limit_type)
        clean_limit_type = re.sub(r"\s+", " ", clean_limit_type).strip()
        # Lone "CREDIT" (after stripping "CASH" via CRASH artefact) → restore full name
        if clean_limit_type == "CREDIT":
            clean_limit_type = "CASH CREDIT"

        # Int rate and Over rate: small floats on secondary row
        small_floats = [(x, v) for x, v in sec_tok_texts
                        if _FLOAT_PAT.match(v) and "." in v and float(v) < 100]
        int_rate  = small_floats[0][1] if len(small_floats) > 0 else ""
        over_rate = small_floats[1][1] if len(small_floats) > 1 else ""

        # Security value amount on secondary row (larger float, x around 428)
        sec_amounts = [(x, v) for x, v in sec_tok_texts if _AMOUNT.match(v)]
        sec_value = sec_amounts[0][1] if sec_amounts else ""

        # Entered by / Verified by: short alpha tokens at far right
        staff_toks = [(x, v) for x, v in sec_tok_texts
                      if re.match(r"^[a-z]{2,5}$", v, re.IGNORECASE) and x > 450]
        entered_by  = staff_toks[0][1] if len(staff_toks) > 0 else ""
        verified_by = staff_toks[1][1] if len(staff_toks) > 1 else ""

        rows.append({
            "Loan Type Code": ltc,
            "Loan Type Name": ltn,
            "A/c No.": acct_no,
            "Name": name,
            "Status": status,
            "Sanction Date": sanction_date,
            "Due Date": due_date,
            "Limit Type": clean_limit_type.strip(),
            "Int. Rate (%)": int_rate,
            "Over Rate (%)": over_rate,
            "Limit Amount (INR)": limit_amount,
            "Security Value Amt (INR)": sec_value,
            "Entered By": entered_by,
            "Verified By": verified_by,
            "Remarks": "",
        })

    # Append summary rows after all account rows
    rows.extend(summary_rows_local)

    df = pd.DataFrame(rows, columns=[
        "Loan Type Code", "Loan Type Name",
        "A/c No.", "Name", "Status",
        "Sanction Date", "Due Date", "Limit Type",
        "Int. Rate (%)", "Over Rate (%)",
        "Limit Amount (INR)", "Security Value Amt (INR)",
        "Entered By", "Verified By", "Remarks",
    ])
    df.attrs["report_date"] = f"{report_date_from} to {report_date_to}"
    return df


# ─────────────────────────────────────────────────────────────────────────────
# 5. TRIAL BALANCE (GL Trial Balance + P&L Ledger)
# Structure: two-column layout — Liabilities (left) | Assets (right)
# Each data row has: [ac_no] [head_name] [balance] on left AND right side
# ─────────────────────────────────────────────────────────────────────────────

# X-coordinate thresholds for Trial Balance two-column layout
_TB_LEFT_ACNO_X  = (20, 50)    # left-side account number
_TB_RIGHT_ACNO_X = (395, 420)  # right-side account number
_TB_LEFT_AMT_X   = (240, 360)  # left-side balance amount
_TB_RIGHT_AMT_X  = (620, 780)  # right-side balance amount


def _in_range(x, rng):
    return rng[0] <= x <= rng[1]


def extract_trial_balance(pdf_path):
    """
    Parses General Ledger/Trial Balance Report.
    Produces two sub-tables in separate sections:
      - GL Trial Balance (Page 1): Liabilities vs Assets
      - Profit & Loss Ledger (Page 2): Income vs Expenditure
    Returns a combined DataFrame with a 'Side' column (LIABILITIES/ASSETS/INCOME/EXPENDITURE).
    """
    rows = []
    report_date = ""

    with pdfplumber.open(pdf_path) as pdf:
        pages = pdf.pages

    # Map page index → section type
    page_sections = {0: ("LIABILITIES", "ASSETS"), 1: ("INCOME", "EXPENDITURE")}

    with pdfplumber.open(pdf_path) as pdf:
        for pi, page in enumerate(pdf.pages):
            left_side, right_side = page_sections.get(pi, ("LEFT", "RIGHT"))
            char_rows = _page_rows(page)

            current_section_left  = ""
            current_section_right = ""

            for top, tokens in char_rows:
                line = _token_text(tokens)

                if _is_noise(line):
                    continue

                # Report date
                m = re.search(r"As on\s*:\s*(\d{2}/\d{2}/\d{4})", line, re.IGNORECASE)
                if m and not report_date:
                    report_date = m.group(1)
                    continue

                # Net Profit line
                m = re.search(r"Net Profit:\s*([\d,\.]+)", line)
                if m:
                    rows.append({
                        "Page": pi + 1, "Side": "NET PROFIT",
                        "Section": "", "A/c No.": "",
                        "Head": "Net Profit",
                        "Balance (INR)": m.group(1),
                        "Is Total": False,
                    })
                    continue

                # Grand total lines (single row with two totals)
                if re.match(r"^\d{3,}[\d,\.]+\s+\d{3,}", line) or \
                   re.match(r"^\d+[\d,\.]+\s+GL TRIAL", line):
                    continue

                # Section headers: "DEPOSITS", "ADVANCES", "INCOME", etc.
                # These appear on left OR right side
                left_section_toks = [t for t in tokens if _in_range(t[0], (60, 210))]
                right_section_toks = [t for t in tokens if _in_range(t[0], (440, 620))]

                # Check for section header (no account number, no amounts)
                has_left_acno  = any(_in_range(t[0], _TB_LEFT_ACNO_X)  and _ACCT6.match(t[2]) for t in tokens)
                has_right_acno = any(_in_range(t[0], _TB_RIGHT_ACNO_X) and _ACCT6.match(t[2]) for t in tokens)
                has_left_amt   = any(_in_range(t[0], _TB_LEFT_AMT_X)  and _AMOUNT.match(t[2]) for t in tokens)
                has_right_amt  = any(_in_range(t[0], _TB_RIGHT_AMT_X) and _AMOUNT.match(t[2]) for t in tokens)

                is_total_row = "****TOTAL****" in line

                if is_total_row:
                    # Extract left and right totals
                    left_amts  = [t[2] for t in tokens if _in_range(t[0], (240, 380)) and _AMOUNT.match(t[2])]
                    right_amts = [t[2] for t in tokens if _in_range(t[0], (620, 780)) and _AMOUNT.match(t[2])]

                    if left_amts:
                        rows.append({
                            "Page": pi + 1, "Side": left_side,
                            "Section": current_section_left, "A/c No.": "****TOTAL****",
                            "Head": "****TOTAL****",
                            "Balance (INR)": left_amts[-1],
                            "Is Total": True,
                        })
                    if right_amts:
                        rows.append({
                            "Page": pi + 1, "Side": right_side,
                            "Section": current_section_right, "A/c No.": "****TOTAL****",
                            "Head": "****TOTAL****",
                            "Balance (INR)": right_amts[-1],
                            "Is Total": True,
                        })
                    continue

                # Section name rows (no acno, no amount) — left side
                if not has_left_acno and not has_left_amt and left_section_toks:
                    left_txt = " ".join(t[2] for t in left_section_toks
                                        if t[2] not in ("****TOTAL****",)
                                        and not _AMOUNT.match(t[2]))
                    if left_txt and not any(c.isdigit() for c in left_txt[:3]):
                        current_section_left = left_txt

                # Section name rows — right side
                if not has_right_acno and not has_right_amt and right_section_toks:
                    right_txt = " ".join(t[2] for t in right_section_toks
                                         if not _AMOUNT.match(t[2]) and t[2] != "****TOTAL****")
                    if right_txt and not any(c.isdigit() for c in right_txt[:3]):
                        current_section_right = right_txt

                # Data rows with account numbers
                if has_left_acno:
                    acno_tok  = next(t for t in tokens if _in_range(t[0], _TB_LEFT_ACNO_X) and _ACCT6.match(t[2]))
                    left_amts = [t[2] for t in tokens if _in_range(t[0], _TB_LEFT_AMT_X) and _AMOUNT.match(t[2])]
                    head_toks = [t[2] for t in tokens
                                 if t[0] > acno_tok[0] + 20 and t[0] < 250
                                 and not _AMOUNT.match(t[2]) and not _ACCT6.match(t[2])]
                    rows.append({
                        "Page": pi + 1, "Side": left_side,
                        "Section": current_section_left,
                        "A/c No.": acno_tok[2],
                        "Head": " ".join(head_toks),
                        "Balance (INR)": left_amts[0] if left_amts else "",
                        "Is Total": False,
                    })

                if has_right_acno:
                    acno_tok   = next(t for t in tokens if _in_range(t[0], _TB_RIGHT_ACNO_X) and _ACCT6.match(t[2]))
                    right_amts = [t[2] for t in tokens if _in_range(t[0], _TB_RIGHT_AMT_X) and _AMOUNT.match(t[2])]
                    head_toks  = [t[2] for t in tokens
                                  if t[0] > acno_tok[0] + 20 and t[0] < 630
                                  and not _AMOUNT.match(t[2]) and not _ACCT6.match(t[2])
                                  and not _in_range(t[0], _TB_LEFT_ACNO_X)]
                    rows.append({
                        "Page": pi + 1, "Side": right_side,
                        "Section": current_section_right,
                        "A/c No.": acno_tok[2],
                        "Head": " ".join(head_toks),
                        "Balance (INR)": right_amts[0] if right_amts else "",
                        "Is Total": False,
                    })

                # Handle rows where left total appears alongside right data
                # (e.g. "****TOTAL****  190998410.36  000229  BUSINESS LOAN...")
                # Already handled above in total_row block.

    df = pd.DataFrame(rows, columns=[
        "Page", "Side", "Section", "A/c No.", "Head", "Balance (INR)", "Is Total"
    ])
    df.attrs["report_date"] = report_date
    return df


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL WRITER
# ─────────────────────────────────────────────────────────────────────────────

_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
_SUMMARY_FILL = PatternFill("solid", fgColor="FFF2CC")
_SUMMARY_FONT = Font(bold=True, name="Arial", size=9, color="1F4E79")
_TOTAL_FILL   = PatternFill("solid", fgColor="D6E4F0")
_TOTAL_FONT   = Font(bold=True, name="Arial", size=9, color="1F4E79")
_SECTION_FILL = PatternFill("solid", fgColor="EBF3FB")
_DATA_FONT    = Font(name="Arial", size=9)
_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center")


def _write_sheet(ws, df, title, report_date=""):
    header_text = f"{title}  |  Period: {report_date}" if report_date else title

    ws.append([header_text])
    cell = ws.cell(row=1, column=1)
    cell.font = Font(bold=True, name="Arial", size=12, color="1F4E79")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    ws.row_dimensions[1].height = 20

    ws.append([])

    ws.append(list(df.columns))
    ws.row_dimensions[3].height = 28
    for col_idx in range(1, len(df.columns) + 1):
        c = ws.cell(row=3, column=col_idx)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = _CENTER
        c.border = _BORDER

    for r_idx, row in enumerate(df.itertuples(index=False), start=4):
        # Determine row style
        remarks = str(getattr(row, "Remarks", "")).strip()
        is_summary = remarks in ("SUMMARY", "SUMMARY ROW")

        acno_val = str(getattr(row, "_3", "")).strip() if hasattr(row, "_3") else ""
        is_total = str(getattr(row, "Is_Total", "")).lower() == "true" or \
                   "****TOTAL****" in str(row)
        is_section_label = any(
            kw in str(row[2]) for kw in
            ["Type Wise", "Branch Wise", "Grand Total", "Typewise",
             "Branchwise", "TOTAL", "Total for"]
        )
        side_val = str(getattr(row, "Side", "")).strip() if hasattr(row, "Side") else ""
        is_side_header = side_val in ("LIABILITIES", "ASSETS", "INCOME", "EXPENDITURE", "NET PROFIT")

        for c_idx, value in enumerate(row, start=1):
            display = "" if str(value) in ("False", "True", "nan") else str(value)
            cell = ws.cell(row=r_idx, column=c_idx, value=display)
            cell.border = _BORDER

            if is_total:
                cell.fill = _TOTAL_FILL
                cell.font = _TOTAL_FONT
                cell.alignment = _CENTER
            elif is_summary or is_section_label:
                cell.fill = _SUMMARY_FILL
                cell.font = _SUMMARY_FONT
                cell.alignment = _CENTER
            elif is_side_header:
                cell.fill = _SECTION_FILL
                cell.font = Font(bold=True, name="Arial", size=9, color="1F4E79")
                cell.alignment = _LEFT
            else:
                cell.font = _DATA_FONT
                cell.alignment = _LEFT

    for col_idx, col_name in enumerate(df.columns, start=1):
        col_vals = df.iloc[:, col_idx - 1].astype(str)
        max_len = max(len(str(col_name)), col_vals.map(len).max() if not df.empty else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A4"


def generate_master_excel(dfs_config, output_path):
    """
    dfs_config: list of (sheet_name, df, title)
    """
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, df, title in dfs_config:
        ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, df, title, df.attrs.get("report_date", ""))

    wb.save(output_path)
    print(f"Saved → {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    UPLOAD_DIR = "/uploads"
    OUTPUT_DIR = "outputs"
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    FILES = {
        "cash":      "D:\\audit_report_generator\\uploads\\PANCHOT BR CASH FEB SUM.pdf",
        "insurance": "D:\\audit_report_generator\\uploads\\PANCHOT BR JAN 2026 INSURANCE PENDING REGI.pdf",
        "overdue":   "D:\\audit_report_generator\\uploads\\PANCHOT BR JAN 2026 OVER DUE REPORT.pdf",
        "limit":     "D:\\audit_report_generator\\uploads\\PANCHOT BR JAN 2026 LIMIT REGI.pdf",
        "trial":     "D:\\audit_report_generator\\uploads\\PANCHOT BR JAN 2026 TRIAL BALANCE.pdf",
    }

    extractors = [
        ("cash",      extract_cash_summary,      "Cash Summary",     "Cash Summary Report – PANCHOT BRANCH"),
        ("insurance", extract_insurance_pending,  "Insurance Pending","Insurance Pending Register – PANCHOT BRANCH"),
        ("overdue",   extract_overdue_report,     "Overdue Report",   "Overdue Report For All – PANCHOT BRANCH"),
        ("limit",     extract_limit_register,     "Limit Register",   "Limit Register – PANCHOT BRANCH"),
        ("trial",     extract_trial_balance,      "Trial Balance",    "GL Trial Balance & P&L Ledger – PANCHOT BRANCH"),
    ]

    dfs_config = []
    for key, extractor_fn, sheet_name, title in extractors:
        fpath = os.path.join(UPLOAD_DIR, FILES[key])
        print(f"Extracting {sheet_name}...")
        df = extractor_fn(fpath)
        print(f"  → {len(df)} rows")
        dfs_config.append((sheet_name, df, title))

    output_path = os.path.join(OUTPUT_DIR, "PANCHOT_MASTER_REPORT.xlsx")
    print(f"\nGenerating Excel workbook → {output_path}")
    generate_master_excel(dfs_config, output_path)

    for sheet_name, df, _ in dfs_config:
        print(f"\n── {sheet_name} ─────────────────────────────")
        print(df.to_string(index=False))
